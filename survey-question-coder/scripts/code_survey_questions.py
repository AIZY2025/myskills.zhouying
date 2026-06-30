#!/usr/bin/env python3
"""
问卷原始数据 → 题号整理

功能:
  从「原始」sheet 读取所有列,在最左侧插入「整理」标识列(值=1),
  在第 1 行写入题号(短码)。支持以下题号规则:
    1) 表头含【XX】(如 2.【G1】、13.【K1】)→ 题号 = XX
    2) 表头是 "数字.中文" 格式(无【】)→ 题号 = Q+数字
    3) 其它(无题号的元数据列)→ 题号 = Q+递增序号(从 max(已有Q+数字)+1 起)
    4) 同一题号对应多列 → 加 -1/-2/-3... 后缀(单列不加)

用法:
  python code_survey_questions.py <input.xlsx> [output.xlsx]
    input.xlsx  : 含「原始」sheet 的工作簿
    output.xlsx : 可选,默认在 input.xlsx 同目录下生成 <原名>_题号化.xlsx

依赖:
  pip install openpyxl
"""
from __future__ import annotations

import argparse
import re
import sys
from collections import OrderedDict
from pathlib import Path

import openpyxl

# 默认不参与"题号"流程的 sheet 名(若工作簿中只有"原始"也可不指定)
DEFAULT_SOURCE_SHEET = "原始"


def extract_base_code(header: str | None, col_idx: int) -> tuple[str, str]:
    """
    从表头中提取题号 base。

    返回 (mode, base):
      mode: 'B' = 【XX】 格式(优先)
            'N' = "数字.中文" 格式
            'X' = 其它/无题号
      base: 题号字符串(不含 -N 后缀);若 header 为空,base 为 None
    """
    if not header:
        return "X", None
    s = str(header).strip()
    # 模式 1:含【XX】(如 "2.【G1】xxx")
    m = re.match(r"^\d+\.【([A-Z]\d+)】", s)
    if m:
        return "B", m.group(1)
    # 模式 2:数字.中文(如 "1.第一部分" "15.您还有...")
    m = re.match(r"^(\d+)\.", s)
    if m:
        return "N", m.group(1)
    # 模式 3:无任何题号(如 "编号"、"语言"、"IP")
    return "X", None


def build_question_codes(ws) -> list[tuple[int, str | None, str]]:
    """
    扫描 ws 的所有列,返回 [(col_idx, header, final_code), ...]。

    final_code 已含 -N 后缀(若该题对应多列)。
    """
    raw_headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    # 第 1 步:收集所有"数字.中文"模式占用的 Q+数字
    used_nums: set[int] = set()
    for h in raw_headers:
        mode, base = extract_base_code(h, 0)
        if mode == "N" and base is not None:
            used_nums.add(int(base))
    max_used = max(used_nums) if used_nums else 0

    # 第 2 步:为每一列生成 base code
    # X 模式从 max_used + 1 开始递增,避免与 N 模式冲突
    items: list[tuple[int, str | None, str, str | None]] = []
    running = max_used
    for c, h in enumerate(raw_headers, 1):
        mode, base = extract_base_code(h, c)
        if mode == "B":
            items.append((c, h, "B", base))
        elif mode == "N":
            items.append((c, h, "N", f"Q{base}"))
        else:  # 'X' 或空
            if h is None:
                # 完全空列:跳过(不参与编码)
                items.append((c, None, "SKIP", None))
            else:
                running += 1
                items.append((c, h, "X", f"Q{running}"))

    # 第 3 步:按 base 分组,同 base 多列加 -N 后缀
    groups: "OrderedDict[str, list[tuple[int, str | None]]]" = OrderedDict()
    for c, h, mode, base in items:
        if mode == "SKIP":
            continue
        groups.setdefault(base, []).append((c, h))

    final: list[tuple[int, str | None, str]] = []
    for base, cols in groups.items():
        if len(cols) == 1:
            final.append((cols[0][0], cols[0][1], base))
        else:
            for i, (c, h) in enumerate(cols, 1):
                final.append((c, h, f"{base}-{i}"))

    return final


def write_coded_sheet(
    src_ws,
    out_wb,
    out_sheet_name: str = "整理",
    mark_col_header: str = "Q",
    mark_col_label: str = "整理",
    mark_col_value=1,
) -> openpyxl.worksheet.worksheet.Worksheet:
    """
    从 src_ws 复制所有数据到 out_wb 的新 sheet out_sheet_name。

    输出表的结构(参考"整理"sheet):
      A1 = mark_col_header     (默认 "Q",作为题号列的标题)
      A2 = mark_col_label      (默认 "整理",作为标识列的列名)
      A3.. = mark_col_value    (默认 1,标识列的数据)

      B1.. = 题号(G1/G1-1/H3-1/K1-2/Q15 等)
      B2.. = 原表头
      B3.. = 原数据(行 1 起的所有数据)
    """
    if out_sheet_name in out_wb.sheetnames:
        del out_wb[out_sheet_name]
    out_ws = out_wb.create_sheet(out_sheet_name)

    final = build_question_codes(src_ws)
    code_by_col = {c: code for c, _, code in final}
    src_cols = sorted(code_by_col.keys())

    # A 列:标识列
    out_ws.cell(1, 1, mark_col_header)   # A1: Q
    out_ws.cell(2, 1, mark_col_label)    # A2: 整理
    for r in range(3, src_ws.max_row + 2):
        out_ws.cell(r, 1, mark_col_value) # A3..: 1

    # B 列起:题号 + 原表头 + 原数据
    # 原表:  r=1        是表头
    #        r=2..N     是数据
    # 整理表: r=1        是题号
    #        r=2        是原表头
    #        r=3..N+1   是原数据
    for new_idx, src_col in enumerate(src_cols, start=2):
        out_ws.cell(1, new_idx, code_by_col[src_col])              # 题号
        out_ws.cell(2, new_idx, src_ws.cell(1, src_col).value)     # 原表头
        for r in range(2, src_ws.max_row + 1):                     # 数据(原 r=2..N)
            out_ws.cell(r + 1, new_idx, src_ws.cell(r, src_col).value)

    return out_ws


def main() -> int:
    p = argparse.ArgumentParser(description="为问卷原始数据每个题目编题号")
    p.add_argument("input", help="输入 .xlsx(必须含「原始」sheet)")
    p.add_argument(
        "-o",
        "--output",
        help="输出 .xlsx 路径(默认: <输入目录>/<原名>_题号化.xlsx)",
    )
    p.add_argument(
        "--source-sheet",
        default=DEFAULT_SOURCE_SHEET,
        help=f"原始 sheet 名(默认: {DEFAULT_SOURCE_SHEET})",
    )
    p.add_argument(
        "--out-sheet",
        default="整理",
        help="输出 sheet 名(默认: 整理)",
    )
    p.add_argument(
        "--mark-col-header",
        default="Q",
        help="A1 标题(默认: Q)",
    )
    p.add_argument(
        "--mark-col-label",
        default="整理",
        help='A2 列名(默认: "整理")',
    )
    p.add_argument(
        "--mark-col-value",
        default=1,
        help="A 列除标题/列名外的数据填充值(默认: 1)",
    )
    args = p.parse_args()

    in_path = Path(args.input).expanduser().resolve()
    if not in_path.exists():
        print(f"❌ 输入文件不存在: {in_path}", file=sys.stderr)
        return 1

    out_path = (
        Path(args.output).expanduser().resolve()
        if args.output
        else in_path.with_name(f"{in_path.stem}_题号化.xlsx")
    )

    wb = openpyxl.load_workbook(in_path, data_only=False)
    if args.source_sheet not in wb.sheetnames:
        print(
            f"❌ 工作簿中没有「{args.source_sheet}」sheet。当前 sheets: {wb.sheetnames}",
            file=sys.stderr,
        )
        return 1

    src_ws = wb[args.source_sheet]
    print(f"📖 读取「{args.source_sheet}」: {src_ws.max_row} 行 × {src_ws.max_column} 列")

    out_ws = write_coded_sheet(
        src_ws,
        wb,
        out_sheet_name=args.out_sheet,
        mark_col_header=args.mark_col_header,
        mark_col_label=args.mark_col_label,
        mark_col_value=args.mark_col_value,
    )

    # 把"整理"表放到最前面(便于查看)
    wb.move_sheet(out_ws, offset=-(wb.sheetnames.index(out_ws.title)))

    wb.save(out_path)
    print(f"✅ 已生成: {out_path}")
    print(f"   整理表: {out_ws.max_row} 行 × {out_ws.max_column} 列")
    print("   题号预览(前 15 列):")
    for c in range(1, min(out_ws.max_column + 1, 16)):
        code = out_ws.cell(1, c).value
        print(f"     {openpyxl.utils.get_column_letter(c)}1: {code!r}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
