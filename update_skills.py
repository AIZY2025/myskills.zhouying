import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
from openpyxl.worksheet.hyperlink import Hyperlink

def find_skill_file(skill_name, base_dir):
    search_folders = ['AI-Skills Directory', 'skillsmp']
    
    for folder_name in search_folders:
        folder_path = os.path.join(base_dir, folder_name)
        if not os.path.exists(folder_path):
            continue
        
        for file in os.listdir(folder_path):
            file_path = os.path.join(folder_path, file)
            if os.path.isfile(file_path):
                file_name_without_ext = os.path.splitext(file)[0]
                
                if file_name_without_ext == skill_name or file == skill_name:
                    file_ext = os.path.splitext(file)[1]
                    file_type = file_ext[1:] if file_ext else ''
                    
                    absolute_path = os.path.abspath(file_path)
                    hyperlink = absolute_path
                    
                    return folder_name, file_type, hyperlink
    
    return '', '', ''

def update_skills_excel():
    excel_path = '/Users/admin/Documents/Skills/Skills.xlsx'
    
    wb = load_workbook(excel_path)
    sheet = wb.active
    
    base_dir = '/Users/admin/Documents/Skills'
    
    for row in range(2, sheet.max_row + 1):
        skill_name = sheet[f'B{row}'].value
        
        if not skill_name:
            continue
        
        folder_name, file_type, hyperlink = find_skill_file(skill_name, base_dir)
        
        sheet[f'E{row}'].value = folder_name
        sheet[f'F{row}'].value = file_type
        
        if hyperlink:
            sheet[f'G{row}'].hyperlink = Hyperlink(target=hyperlink, display=hyperlink, ref=f'G{row}')
        else:
            sheet[f'G{row}'].value = ''
            sheet[f'G{row}'].hyperlink = None
        
        print(f"处理: {skill_name} -> 文件夹: {folder_name}, 类型: {file_type}")
    
    for column in sheet.columns:
        sheet.column_dimensions[column[0].column_letter].width = 15
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='left')
            if cell.value is not None and cell.value != '':
                cell.border = thin_border
    
    sheet.freeze_panes = 'A2'
    sheet.auto_filter.ref = sheet.dimensions
    
    wb.save(excel_path)
    print(f"\nExcel文件已更新并设置格式: {excel_path}")

if __name__ == '__main__':
    update_skills_excel()
